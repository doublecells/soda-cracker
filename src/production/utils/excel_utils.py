
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

class ExcelUtils:
    """
    A class processing Microsft Excel file
    """

    def __init__(self, entity) -> None:
        self._entity = entity
        self._wb = None
        self._ws = None
        self._range = None
        self._data = list()

    def get_columns(self, row_num = 1):
        """
        Read data from sepcific row as column row
        """
        self._internal_init()

        for col in self.get_rows(row_num, row_num):
            return self._get_cell_values(col)

    def get_rows(self, begin_row = 1, end_row = None):
        """        
        Read data from Excel file with all activate rows or specific rows
        """
        self._internal_init()

        _row_data = list()
        for row in self._range.iter_rows(min_row=begin_row, max_row=end_row):
            _row_data.append(row)
        return _row_data

    def create(self):
        """
        Generate file with properties
        """
        self._wb = Workbook()
        if self._entity.worksheet is not None:
            self._ws = self._wb.active
            self._ws.title = self._entity.worksheet

        self._wb.save(self._entity.path)

    def save_columns(self, columns):
        """
        Write data from array data as column names
        """
        self._internal_save_init()

        self._ws = self._wb[self._entity.worksheet]
        
        _col_num = 1

        for col in columns:
            self._ws.cell(1, _col_num).value = col
            self._ws.column_dimensions[get_column_letter(_col_num)].width = len(col) + 1
            _col_num += 1

        self._wb.save(self._entity.path)

    def save_rows(self, rows):
        """
        Write data from array data as row values
        """
        self._internal_save_init()
        self._ws = self._wb[self._entity.worksheet]

        _row_num = 2
        for row in rows:
            self._save_cell_values(_row_num, row)
            _row_num += 1
        
        self._wb.save(self._entity.path)
    
    def _save_cell_values(self, row_num=2, args=[]):
        for idx in range(0, len(args)):
            self._ws.cell(row_num, idx + 1).value = args[idx]


    def _get_cell_values(self, args):
        _cell_values = list()
        for idx in range(0, len(args)):
            _cell_values.append(args[idx].value)
        return _cell_values

    def _print_cell_value(self, args):
        """
        Print cell value in each of cells
        """
        for idx in range(0, len(args)):
            print(args[idx].value)

    def _internal_init(self) -> None:
        self._wb = load_workbook(filename=self._entity.path)
        if self._entity.worksheet is not None:
            self._range = self._wb[self._entity.worksheet]
        else:
            self._range = self._wb.worksheets[self._entity.index]

    def _internal_save_init(self) -> None:
        self._wb = load_workbook(filename=self._entity.path)
        self._ws = self._wb[self._entity.worksheet]        
